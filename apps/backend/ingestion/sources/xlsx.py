from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from connection._shared import is_empty_row, normalize_header, normalize_preview_row


@dataclass
class SheetData:
    sheet_name: str
    rows: list[tuple | list]
    is_visible: bool = True
    row_count: int = 0
    column_count: int = 0
    merged_cell_ranges: list[str] = field(default_factory=list)
    contains_formulas: bool = False
    warnings: list[str] = field(default_factory=list)
    raw_cells: list | None = None


@dataclass
class WorkbookData:
    sheets: list[SheetData]
    metadata: dict = field(default_factory=dict)


def sample_rows(rows, limit: int):
    return list(rows[:limit])


def _propagate_merged_headers(rows, merged_ranges):
    if not rows:
        return rows
    output = [list(row) for row in rows]
    for merge_range in merged_ranges:
        try:
            start, end = merge_range.split(":")
            start_col = "".join(ch for ch in start if ch.isalpha())
            start_row = int("".join(ch for ch in start if ch.isdigit()))
            end_col = "".join(ch for ch in end if ch.isalpha())
            end_row = int("".join(ch for ch in end if ch.isdigit()))
            if start_row != end_row:
                continue
            row_index = start_row - 1
            if row_index >= len(output):
                continue
            start_idx = _col_to_index(start_col)
            end_idx = _col_to_index(end_col)
            value = output[row_index][start_idx] if start_idx < len(output[row_index]) else None
            for idx in range(start_idx, end_idx + 1):
                while idx >= len(output[row_index]):
                    output[row_index].append(None)
                if idx != start_idx:
                    output[row_index][idx] = value
        except Exception:
            continue
    return [tuple(row) for row in output]


def _col_to_index(col: str) -> int:
    value = 0
    for char in col.upper():
        value = value * 26 + (ord(char) - 64)
    return value - 1


def _suggest_alternate_header(rows) -> bool:
    if len(rows) < 3:
        return False
    first = rows[0]
    second = rows[1]
    third = rows[2]
    first_strings = all(isinstance(v, str) for v in first if v is not None)
    second_strings = all(isinstance(v, str) for v in second if v is not None)
    third_has_data = any(not isinstance(v, str) for v in third if v is not None)
    return first_strings and second_strings and third_has_data


def _read_rows_from_worksheet(worksheet):
    merged_ranges = [str(merged) for merged in worksheet.merged_cells.ranges]
    raw_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    propagated_rows = _propagate_merged_headers(raw_rows, merged_ranges)
    return propagated_rows, merged_ranges


def read_workbook(storage_path: Path):
    try:
        workbook = load_workbook(filename=str(storage_path), read_only=False, data_only=False)
    except Exception:
        return []
    try:
        results = []
        for worksheet in workbook.worksheets:
            rows, merged_ranges = _read_rows_from_worksheet(worksheet)
            non_empty_rows = [row for row in rows if not is_empty_row(row)]
            header_row = non_empty_rows[0] if non_empty_rows else []
            preview_columns = normalize_header(list(header_row), len(header_row)) if header_row else []
            preview_rows = [normalize_preview_row(list(row), len(preview_columns)) for row in non_empty_rows[1:11]]
            contains_formulas = any(
                isinstance(cell.value, str) and cell.value.startswith("=")
                for row in worksheet.iter_rows()
                for cell in row
            )
            results.append(
                {
                    "sheet_name": worksheet.title,
                    "rows": [tuple(row) for row in non_empty_rows],
                    "is_visible": worksheet.sheet_state == "visible",
                    "row_count": len(non_empty_rows),
                    "data_row_count": max(len(non_empty_rows) - (1 if preview_columns else 0), 0),
                    "column_count": max((len(row) for row in non_empty_rows), default=0),
                    "header_row_index": 0 if preview_columns else None,
                    "confidence": 1.0 if preview_columns else 0.0,
                    "warnings": [] if non_empty_rows else ["Empty sheet"],
                    "merged_cell_ranges": merged_ranges,
                    "contains_formulas": contains_formulas,
                    "preview_columns": preview_columns,
                    "preview_rows": preview_rows,
                    "raw_cells": rows,
                }
            )
        return results
    finally:
        workbook.close()


def get_workbook_metadata(storage_path: Path) -> dict:
    if not storage_path.exists():
        return {"error": "File not found"}
    try:
        workbook = load_workbook(filename=str(storage_path), read_only=False, data_only=False)
    except Exception:
        return {"error": "Invalid workbook"}
    try:
        sheets = read_workbook(storage_path)
        visible_count = sum(1 for sheet in sheets if sheet["is_visible"])
        return {
            "sheet_count": len(sheets),
            "visible_sheet_count": visible_count,
            "hidden_sheet_count": len(sheets) - visible_count,
            "contains_formulas": any(sheet["contains_formulas"] for sheet in sheets),
        }
    finally:
        workbook.close()


def get_sheet_preview(storage_path: Path, sheet_name: str, max_rows: int = 10):
    sheets = read_workbook(storage_path)
    for sheet in sheets:
        if sheet["sheet_name"] == sheet_name:
            return [list(row) for row in sheet["rows"][:max_rows]]
    return []

