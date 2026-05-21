from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from common.errors import ValidationError
from connection._shared import (
    is_empty_row,
    normalize_header,
    write_jsonl_version,
)


def materialize_dataset_version(storage_path: Path, sheet_name: str, dataset_id: str) -> tuple[Path, int, int]:
    if storage_path.suffix.lower() == ".csv":
        return _materialize_csv_version(storage_path, dataset_id)

    workbook = load_workbook(filename=str(storage_path), read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValidationError(f"Sheet '{sheet_name}' not found in workbook")

        worksheet = workbook[sheet_name]
        rows = [list(row) for row in worksheet.iter_rows(values_only=True) if not is_empty_row(row)]
        if not rows:
            return write_jsonl_version([], dataset_id, sheet_name), 0, 0

        header_row = rows[0]
        column_count = len(header_row)
        headers = normalize_header(header_row, column_count)
        output_rows = [
            {header: row[index] if index < len(row) else None for index, header in enumerate(headers)}
            for row in rows[1:]
        ]

        return write_jsonl_version(output_rows, dataset_id, sheet_name), len(output_rows), column_count
    finally:
        workbook.close()


def _materialize_csv_version(storage_path: Path, dataset_id: str) -> tuple[Path, int, int]:
    with open(storage_path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        rows = [row for row in reader if not is_empty_row(row)]

    if not rows:
        return write_jsonl_version([], dataset_id, storage_path.stem), 0, 0

    headers = normalize_header(rows[0], len(rows[0]))
    output_rows = []
    for row in rows[1:]:
        output_rows.append({header: row[index] if index < len(row) else None for index, header in enumerate(headers)})

    return write_jsonl_version(output_rows, dataset_id, storage_path.stem), len(output_rows), len(headers)
