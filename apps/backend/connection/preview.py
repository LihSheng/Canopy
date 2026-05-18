from __future__ import annotations

import csv
import uuid
from pathlib import Path

from fastapi import UploadFile
from openpyxl import load_workbook

from common.errors import ValidationError
from connection._shared import (
    is_empty_row,
    normalize_header,
    normalize_preview_row,
    storage_root,
)


def save_uploaded_file(file: UploadFile) -> Path:
    if file.filename is None:
        raise ValidationError("File name is required")

    suffix = Path(file.filename).suffix.lower()
    if suffix not in {".xlsx", ".csv"}:
        raise ValidationError("Only .xlsx and .csv files are supported for static file sources")

    storage_dir = storage_root() / str(uuid.uuid4())
    storage_dir.mkdir(parents=True, exist_ok=True)

    storage_path = storage_dir / file.filename
    storage_path.write_bytes(file.file.read())
    return storage_path


def delete_uploaded_file(storage_path: Path) -> None:
    resolved_path = storage_path.resolve()
    root = storage_root().resolve()

    if resolved_path != root and root not in resolved_path.parents:
        raise ValidationError("Invalid preview file path")

    if resolved_path.exists():
        resolved_path.unlink()

    current_dir = resolved_path.parent
    while current_dir != root and current_dir.exists():
        try:
            current_dir.rmdir()
        except OSError:
            break
        current_dir = current_dir.parent


def build_sheet_profiles(storage_path: Path) -> list[dict]:
    if storage_path.suffix.lower() == ".csv":
        return [_build_csv_profile(storage_path)]

    workbook = load_workbook(filename=str(storage_path), read_only=True, data_only=True)
    try:
        profiles: list[dict] = []
        for worksheet in workbook.worksheets:
            row_count = 0
            column_count = 0
            header_row_index = None
            preview_columns: list[str] = []
            preview_rows: list[list[object | None]] = []

            for index, row in enumerate(worksheet.iter_rows(values_only=True)):
                if is_empty_row(row):
                    continue
                if header_row_index is None:
                    header_row_index = index
                    column_count = max(column_count, len(row))
                    preview_columns = normalize_header(list(row), len(row))
                    continue
                row_count += 1
                column_count = max(column_count, len(row))
                if len(preview_rows) < 10:
                    preview_rows.append(normalize_preview_row(list(row), len(preview_columns)))

            profiles.append(
                {
                    "sheet_name": worksheet.title,
                    "row_count": row_count + (1 if header_row_index is not None else 0),
                    "data_row_count": row_count,
                    "column_count": column_count,
                    "header_row_index": header_row_index,
                    "confidence": 1.0 if header_row_index is not None else 0.0,
                    "warnings": [] if header_row_index is not None else ["Empty sheet"],
                    "preview_columns": preview_columns,
                    "preview_rows": preview_rows,
                },
            )

        return profiles
    finally:
        workbook.close()


def _build_csv_profile(storage_path: Path) -> dict:
    with open(storage_path, "r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        rows = [row for row in reader if not is_empty_row(row)]

    if not rows:
        return {
            "sheet_name": storage_path.stem,
            "row_count": 0,
            "data_row_count": 0,
            "column_count": 0,
            "header_row_index": None,
            "confidence": 0.0,
            "warnings": ["Empty file"],
            "preview_columns": [],
            "preview_rows": [],
        }

    headers = normalize_header(rows[0], len(rows[0]))
    return {
        "sheet_name": storage_path.stem,
        "row_count": len(rows),
        "data_row_count": max(len(rows) - 1, 0),
        "column_count": len(rows[0]),
        "header_row_index": 0,
        "confidence": 1.0,
        "warnings": [],
        "preview_columns": headers,
        "preview_rows": [normalize_preview_row(row, len(headers)) for row in rows[1:11]],
    }

