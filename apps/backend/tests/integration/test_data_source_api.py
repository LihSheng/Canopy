from __future__ import annotations

import io
import uuid
from datetime import UTC, datetime

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import text

from dataset.domain import Dataset
from dataset.repository import DatasetRepository
from semantic.domain import PropertyMapping, SemanticMapping
from semantic.repository import ObjectTypeRepository, SemanticMappingRepository
from semantic.service import ObjectTypeService

pytestmark = pytest.mark.api_schema


def _make_xlsx_bytes() -> io.BytesIO:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Payroll"
    sheet.append(["name", "amount"])
    sheet.append(["Alice", 100])
    sheet.append(["Bob", 200])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def test_static_file_preview_and_dataset_creation(client: TestClient, auth_headers, monkeypatch, tmp_path):
    from common.config import settings

    monkeypatch.setattr(settings, "export_storage_dir", str(tmp_path), raising=False)

    project_resp = client.post(
        "/api/projects/",
        json={"name": "Import Project", "description": "Test project"},
        headers=auth_headers,
    )
    assert project_resp.status_code == 201
    project_id = project_resp.json()["id"]

    preview_resp = client.post(
        "/api/connections/preview",
        files={
            "file": (
                "payroll.xlsx",
                _make_xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_headers,
    )
    assert preview_resp.status_code == 200
    preview = preview_resp.json()
    assert preview["file_name"] == "payroll.xlsx"
    assert preview["sheet_profiles"][0]["sheet_name"] == "Payroll"
    assert preview["sheet_profiles"][0]["row_count"] == 3
    assert preview["sheet_profiles"][0]["data_row_count"] == 2
    assert preview["sheet_profiles"][0]["preview_columns"] == ["name", "amount"]
    assert preview["sheet_profiles"][0]["preview_rows"][0] == ["Alice", 100]

    connection_resp = client.post(
        "/api/connections/",
        json={
            "project_id": project_id,
            "source_type": "static_file",
            "name": "Payroll file",
            "config_json": {
                "file_name": preview["file_name"],
                "source_file_path": preview["source_file_path"],
            },
        },
        headers=auth_headers,
    )
    assert connection_resp.status_code == 201
    connection_id = connection_resp.json()["id"]

    dataset_resp = client.post(
        "/api/datasets/",
        json={
            "project_id": project_id,
            "connection_id": connection_id,
            "name": "Payroll",
            "source_object_name": "Payroll",
        },
        headers=auth_headers,
    )
    assert dataset_resp.status_code == 201
    dataset = dataset_resp.json()
    assert dataset["active_version_id"] is not None

    versions_resp = client.get(f"/api/datasets/{dataset['id']}/versions", headers=auth_headers)
    assert versions_resp.status_code == 200
    versions = versions_resp.json()
    assert len(versions) == 1
    assert versions[0]["status"] == "ready"
    assert versions[0]["row_count"] == 2

    preview_dataset_resp = client.get(f"/api/datasets/{dataset['id']}/preview", headers=auth_headers)
    assert preview_dataset_resp.status_code == 200
    preview_dataset = preview_dataset_resp.json()
    assert preview_dataset["columns"] == ["name", "amount"]
    assert preview_dataset["rows"][0] == ["Alice", 100]

    delete_preview_resp = client.request(
        "DELETE",
        "/api/connections/preview",
        json={"source_file_path": preview["source_file_path"]},
        headers=auth_headers,
    )
    assert delete_preview_resp.status_code == 200
    assert delete_preview_resp.json()["deleted"] is True


def test_preview_backfills_legacy_dataset_without_active_version(
    client: TestClient, auth_headers, monkeypatch, tmp_path, db_session
):
    from common.config import settings

    monkeypatch.setattr(settings, "export_storage_dir", str(tmp_path), raising=False)

    project_resp = client.post(
        "/api/projects/",
        json={"name": "Legacy Project", "description": "Test project"},
        headers=auth_headers,
    )
    assert project_resp.status_code == 201
    project_id = project_resp.json()["id"]

    preview_resp = client.post(
        "/api/connections/preview",
        files={
            "file": (
                "payroll.xlsx",
                _make_xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_headers,
    )
    assert preview_resp.status_code == 200
    preview = preview_resp.json()

    connection_resp = client.post(
        "/api/connections/",
        json={
            "project_id": project_id,
            "source_type": "static_file",
            "name": "Payroll file",
            "config_json": {
                "file_name": preview["file_name"],
                "source_file_path": preview["source_file_path"],
            },
        },
        headers=auth_headers,
    )
    assert connection_resp.status_code == 201
    connection_id = connection_resp.json()["id"]

    dataset_repo = DatasetRepository(db_session)
    dataset_repo.save(
        Dataset(
            id=str(uuid.uuid4()),
            project_id=project_id,
            connection_id=connection_id,
            name="Payroll Legacy",
            source_object_name="Payroll",
            tenant_id="test-tenant-1",
        ),
    )

    datasets_resp = client.get("/api/datasets/", headers=auth_headers)
    dataset_id = next(item["id"] for item in datasets_resp.json() if item["name"] == "Payroll Legacy")

    preview_dataset_resp = client.get(f"/api/datasets/{dataset_id}/preview", headers=auth_headers)
    assert preview_dataset_resp.status_code == 200
    preview_dataset = preview_dataset_resp.json()
    assert preview_dataset["columns"] == ["name", "amount"]
    assert preview_dataset["rows"][0] == ["Alice", 100]

    versions_resp = client.get(f"/api/datasets/{dataset_id}/versions", headers=auth_headers)
    assert versions_resp.status_code == 200
    versions = versions_resp.json()
    assert len(versions) == 1


def test_preview_backfills_upload_id_connection_without_active_version(
    client: TestClient, auth_headers, monkeypatch, tmp_path, db_session
):
    from common.config import settings

    monkeypatch.setattr(settings, "export_storage_dir", str(tmp_path), raising=False)

    project_resp = client.post(
        "/api/projects/",
        json={"name": "Upload ID Project", "description": "Test project"},
        headers=auth_headers,
    )
    assert project_resp.status_code == 201
    project_id = project_resp.json()["id"]

    preview_resp = client.post(
        "/api/connections/preview",
        files={
            "file": (
                "payroll.xlsx",
                _make_xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_headers,
    )
    assert preview_resp.status_code == 200
    preview = preview_resp.json()
    upload_id = str(uuid.uuid4())

    db_session.execute(
        text(
            """
            create table if not exists uploads (
                id varchar primary key,
                storage_path varchar not null
            )
            """,
        ),
    )
    db_session.execute(
        text("insert into uploads (id, storage_path) values (:id, :storage_path)"),
        {"id": upload_id, "storage_path": preview["source_file_path"]},
    )
    db_session.commit()

    connection_resp = client.post(
        "/api/connections/",
        json={
            "project_id": project_id,
            "source_type": "static_file",
            "name": "Payroll file",
            "config_json": {"upload_id": upload_id},
        },
        headers=auth_headers,
    )
    assert connection_resp.status_code == 201
    connection_id = connection_resp.json()["id"]

    dataset_repo = DatasetRepository(db_session)
    dataset_repo.save(
        Dataset(
            id=str(uuid.uuid4()),
            project_id=project_id,
            connection_id=connection_id,
            name="Payroll Upload Legacy",
            source_object_name="Payroll",
            tenant_id="test-tenant-1",
        ),
    )

    datasets_resp = client.get("/api/datasets/", headers=auth_headers)
    dataset_id = next(item["id"] for item in datasets_resp.json() if item["name"] == "Payroll Upload Legacy")

    preview_dataset_resp = client.get(f"/api/datasets/{dataset_id}/preview", headers=auth_headers)
    assert preview_dataset_resp.status_code == 200
    preview_dataset = preview_dataset_resp.json()
    assert preview_dataset["columns"] == ["name", "amount"]
    assert preview_dataset["rows"][0] == ["Alice", 100]


def _create_dataset_via_api(
    client,
    auth_headers,
    monkeypatch,
    tmp_path,
    xlsx_buffer,
    project_name,
    connection_name,
    dataset_name,
    sheet_name="Payroll",
):
    from common.config import settings

    monkeypatch.setattr(settings, "export_storage_dir", str(tmp_path), raising=False)

    project_resp = client.post(
        "/api/projects/",
        json={"name": project_name, "description": "Test project"},
        headers=auth_headers,
    )
    assert project_resp.status_code == 201
    project_id = project_resp.json()["id"]

    preview_resp = client.post(
        "/api/connections/preview",
        files={"file": ("data.xlsx", xlsx_buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=auth_headers,
    )
    assert preview_resp.status_code == 200
    preview = preview_resp.json()

    connection_resp = client.post(
        "/api/connections/",
        json={
            "project_id": project_id,
            "source_type": "static_file",
            "name": connection_name,
            "config_json": {
                "file_name": preview["file_name"],
                "source_file_path": preview["source_file_path"],
            },
        },
        headers=auth_headers,
    )
    assert connection_resp.status_code == 201
    connection_id = connection_resp.json()["id"]

    dataset_resp = client.post(
        "/api/datasets/",
        json={
            "project_id": project_id,
            "connection_id": connection_id,
            "name": dataset_name,
            "source_object_name": sheet_name,
        },
        headers=auth_headers,
    )
    assert dataset_resp.status_code == 201
    dataset = dataset_resp.json()
    assert dataset["active_version_id"] is not None
    return dataset["id"]


def _make_two_row_xlsx_bytes() -> io.BytesIO:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Payroll"
    sheet.append(["name", "amount"])
    sheet.append(["Alice", 100])
    sheet.append(["Bob", 200])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def test_dataset_preview_pagination_first_page(client: TestClient, auth_headers, monkeypatch, tmp_path):
    dataset_id = _create_dataset_via_api(
        client,
        auth_headers,
        monkeypatch,
        tmp_path,
        _make_two_row_xlsx_bytes(),
        "Pagination Project",
        "Pagination Conn",
        "Paginated Dataset",
    )

    resp = client.get(
        f"/api/datasets/{dataset_id}/preview",
        params={"page": 1, "page_size": 1},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["columns"] == ["name", "amount"]
    assert data["rows"] == [["Alice", 100]]
    assert data["total_row_count"] == 2
    assert data["filtered_row_count"] == 2
    assert data["page"] == 1
    assert data["page_size"] == 1


def test_dataset_preview_pagination_second_page(client: TestClient, auth_headers, monkeypatch, tmp_path):
    dataset_id = _create_dataset_via_api(
        client,
        auth_headers,
        monkeypatch,
        tmp_path,
        _make_two_row_xlsx_bytes(),
        "Pagination2 Project",
        "Pagination2 Conn",
        "Paginated2 Dataset",
    )

    resp = client.get(
        f"/api/datasets/{dataset_id}/preview",
        params={"page": 2, "page_size": 1},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["rows"] == [["Bob", 200]]
    assert data["total_row_count"] == 2
    assert data["filtered_row_count"] == 2
    assert data["page"] == 2
    assert data["page_size"] == 1


def _make_large_xlsx_bytes(num_rows: int = 60) -> io.BytesIO:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Payroll"
    sheet.append(["name", "amount"])
    for i in range(num_rows):
        if i == 42:
            sheet.append(["UniqueSearchTarget", 9999])
        else:
            sheet.append([f"Person_{i}", i * 10])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def test_dataset_preview_search_finds_match(client: TestClient, auth_headers, monkeypatch, tmp_path):
    dataset_id = _create_dataset_via_api(
        client,
        auth_headers,
        monkeypatch,
        tmp_path,
        _make_large_xlsx_bytes(60),
        "Search Project",
        "Search Conn",
        "Search Dataset",
    )

    resp = client.get(
        f"/api/datasets/{dataset_id}/preview",
        params={"search": "UniqueSearchTarget", "page_size": 100},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["total_row_count"] == 60
    assert data["filtered_row_count"] == 1
    assert len(data["rows"]) == 1
    assert data["rows"][0] == ["UniqueSearchTarget", 9999]


def test_dataset_preview_search_no_match(client: TestClient, auth_headers, monkeypatch, tmp_path):
    dataset_id = _create_dataset_via_api(
        client,
        auth_headers,
        monkeypatch,
        tmp_path,
        _make_large_xlsx_bytes(60),
        "SearchNoMatch Project",
        "SearchNoMatch Conn",
        "SearchNoMatch Dataset",
    )

    resp = client.get(
        f"/api/datasets/{dataset_id}/preview",
        params={"search": "zzzzzzNotFound"},
        headers=auth_headers,
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["total_row_count"] == 60
    assert data["filtered_row_count"] == 0
    assert data["rows"] == []


def test_dataset_preview_invalid_page(client: TestClient, auth_headers, monkeypatch, tmp_path):
    dataset_id = _create_dataset_via_api(
        client,
        auth_headers,
        monkeypatch,
        tmp_path,
        _make_two_row_xlsx_bytes(),
        "InvalidPage Project",
        "InvalidPage Conn",
        "InvalidPage Dataset",
    )

    resp = client.get(
        f"/api/datasets/{dataset_id}/preview",
        params={"page": 0},
        headers=auth_headers,
    )
    assert resp.status_code == 422

    resp = client.get(
        f"/api/datasets/{dataset_id}/preview",
        params={"page": -1},
        headers=auth_headers,
    )
    assert resp.status_code == 422


def _create_project_and_connection(
    client: TestClient,
    auth_headers,
    project_name: str = "Lifecycle Project",
):
    project_resp = client.post(
        "/api/projects/",
        json={"name": project_name, "description": "Test project"},
        headers=auth_headers,
    )
    assert project_resp.status_code == 201
    project_id = project_resp.json()["id"]

    connection_resp = client.post(
        "/api/connections/",
        json={
            "project_id": project_id,
            "source_type": "static_file",
            "name": "Lifecycle file",
            "config_json": {"file_name": "source.xlsx"},
        },
        headers=auth_headers,
    )
    assert connection_resp.status_code == 201
    return project_id, connection_resp.json()["id"]


def test_connection_lifecycle_pause_and_audit(client: TestClient, auth_headers, db_session):
    _, connection_id = _create_project_and_connection(client, auth_headers, "Pause Project")

    response = client.post(f"/api/connections/{connection_id}/pause", headers=auth_headers)

    assert response.status_code == 200
    assert response.json()["status"] == "paused"

    audit_event = db_session.execute(
        text(
            """
            select event_type, event_payload_json
            from audit_events
            where event_type = 'connection.paused'
            """
        )
    ).first()
    assert audit_event is not None
    assert connection_id in audit_event.event_payload_json


def test_connection_delete_allowed_when_unused(client: TestClient, auth_headers):
    _, connection_id = _create_project_and_connection(client, auth_headers, "Unused Delete Project")

    delete_resp = client.delete(f"/api/connections/{connection_id}", headers=auth_headers)

    assert delete_resp.status_code == 200
    assert delete_resp.json() == {"deleted": True, "id": connection_id}

    get_resp = client.get(f"/api/connections/{connection_id}", headers=auth_headers)
    assert get_resp.status_code == 404

    list_resp = client.get("/api/connections/", headers=auth_headers)
    assert list_resp.status_code == 200
    assert all(item["id"] != connection_id for item in list_resp.json())


def test_connection_soft_delete_blocks_active_dependencies(client: TestClient, auth_headers):
    project_id, connection_id = _create_project_and_connection(
        client,
        auth_headers,
        "Dependency Project",
    )
    dataset_resp = client.post(
        "/api/datasets/",
        json={
            "project_id": project_id,
            "connection_id": connection_id,
            "name": "Dependent dataset",
            "source_object_name": "Payroll",
        },
        headers=auth_headers,
    )
    assert dataset_resp.status_code == 201

    dependency_resp = client.get(
        f"/api/connections/{connection_id}/dependencies",
        headers=auth_headers,
    )
    assert dependency_resp.status_code == 200
    assert dependency_resp.json()["can_delete"] is False
    assert dependency_resp.json()["active_dataset_count"] == 1

    delete_resp = client.delete(
        f"/api/connections/{connection_id}",
        headers=auth_headers,
    )

    assert delete_resp.status_code == 400
    assert "active dependencies" in delete_resp.json()["detail"]


def test_dataset_delete_allowed_when_unused(client: TestClient, auth_headers, monkeypatch, tmp_path):
    dataset_id = _create_dataset_via_api(
        client,
        auth_headers,
        monkeypatch,
        tmp_path,
        _make_two_row_xlsx_bytes(),
        "Delete Dataset Project",
        "Delete Dataset Conn",
        "Delete Dataset",
    )

    delete_summary_resp = client.get(
        f"/api/datasets/{dataset_id}/dependencies",
        headers=auth_headers,
    )
    assert delete_summary_resp.status_code == 200
    delete_summary = delete_summary_resp.json()
    assert delete_summary["can_delete"] is True
    assert delete_summary["active_run_count"] == 0
    assert delete_summary["entity_count"] == 0

    delete_resp = client.delete(f"/api/datasets/{dataset_id}", headers=auth_headers)
    assert delete_resp.status_code == 200
    assert delete_resp.json() == {"deleted": True, "id": dataset_id}

    dataset_storage_dir = tmp_path / "data-sources" / dataset_id
    assert not dataset_storage_dir.exists()

    get_resp = client.get(f"/api/datasets/{dataset_id}", headers=auth_headers)
    assert get_resp.status_code == 404

    versions_resp = client.get(f"/api/datasets/{dataset_id}/versions", headers=auth_headers)
    assert versions_resp.status_code == 200
    assert versions_resp.json() == []


def test_dataset_delete_blocks_active_runs(client: TestClient, auth_headers, monkeypatch, tmp_path):
    dataset_id = _create_dataset_via_api(
        client,
        auth_headers,
        monkeypatch,
        tmp_path,
        _make_two_row_xlsx_bytes(),
        "Blocked Dataset Project",
        "Blocked Dataset Conn",
        "Blocked Dataset",
    )

    dataset = client.get(f"/api/datasets/{dataset_id}", headers=auth_headers).json()
    run_resp = client.post(
        "/api/runs/",
        json={
            "project_id": dataset["project_id"],
            "connection_id": dataset["connection_id"],
            "dataset_id": dataset_id,
            "started_by": "tester",
        },
        headers=auth_headers,
    )
    assert run_resp.status_code == 201

    delete_summary_resp = client.get(
        f"/api/datasets/{dataset_id}/dependencies",
        headers=auth_headers,
    )
    assert delete_summary_resp.status_code == 200
    delete_summary = delete_summary_resp.json()
    assert delete_summary["can_delete"] is False
    assert delete_summary["active_run_count"] == 1
    assert delete_summary["entity_count"] == 0

    delete_resp = client.delete(f"/api/datasets/{dataset_id}", headers=auth_headers)
    assert delete_resp.status_code == 400
    assert "active run" in delete_resp.json()["detail"].lower()


def test_dataset_delete_blocks_entity_dependencies(client: TestClient, auth_headers, monkeypatch, tmp_path, db_session):
    dataset_id = _create_dataset_via_api(
        client,
        auth_headers,
        monkeypatch,
        tmp_path,
        _make_two_row_xlsx_bytes(),
        "Entity Locked Project",
        "Entity Locked Conn",
        "Entity Locked Dataset",
    )
    active_version_id = client.get(f"/api/datasets/{dataset_id}/versions", headers=auth_headers).json()[0]["id"]

    object_type = ObjectTypeService(ObjectTypeRepository(db_session)).create(
        tenant_id="test-tenant-1",
        object_type_key="leave_request",
        display_name="Leave Request",
        description="Entity using dataset",
    )
    SemanticMappingRepository(db_session).save(
        SemanticMapping(
            id=str(uuid.uuid4()),
            tenant_id="test-tenant-1",
            dataset_id=dataset_id,
            dataset_version_id=active_version_id,
            version_number=1,
            object_type_id=object_type.id,
            object_type_key=object_type.object_type_key,
            properties=[
                PropertyMapping(
                    source_column="id",
                    property_name="id",
                    semantic_type="integer",
                    included=True,
                    is_primary_key=True,
                )
            ],
            links=[],
            computed_properties=[],
            source_nodes=[],
            created_at=datetime.now(UTC),
            updated_at=datetime.now(UTC),
        )
    )

    delete_summary_resp = client.get(
        f"/api/datasets/{dataset_id}/dependencies",
        headers=auth_headers,
    )
    assert delete_summary_resp.status_code == 200
    delete_summary = delete_summary_resp.json()
    assert delete_summary["can_delete"] is False
    assert delete_summary["active_run_count"] == 0
    assert delete_summary["entity_count"] == 1
    assert "used by 1 entity" in delete_summary["blocking_reason"].lower()

    delete_resp = client.delete(f"/api/datasets/{dataset_id}", headers=auth_headers)
    assert delete_resp.status_code == 400
    assert "used by 1 entity" in delete_resp.json()["detail"].lower()


def test_dataset_version_delete_allowed_for_non_active_version(
    client: TestClient,
    auth_headers,
    monkeypatch,
    tmp_path,
):
    dataset_id = _create_dataset_via_api(
        client,
        auth_headers,
        monkeypatch,
        tmp_path,
        _make_two_row_xlsx_bytes(),
        "Version Delete Project",
        "Version Delete Conn",
        "Version Delete Dataset",
    )

    versions_resp = client.get(f"/api/datasets/{dataset_id}/versions", headers=auth_headers)
    versions = versions_resp.json()
    assert len(versions) == 1
    active_version_id = versions[0]["id"]

    second_version_resp = client.post(
        f"/api/datasets/{dataset_id}/versions",
        json={},
        headers=auth_headers,
    )
    assert second_version_resp.status_code == 201
    second_version_id = second_version_resp.json()["id"]
    assert second_version_id != active_version_id

    version_summary_resp = client.get(
        f"/api/datasets/{dataset_id}/versions/{second_version_id}/dependencies",
        headers=auth_headers,
    )
    assert version_summary_resp.status_code == 200
    version_summary = version_summary_resp.json()
    assert version_summary["can_delete"] is True
    assert version_summary["is_active_version"] is False

    delete_resp = client.delete(
        f"/api/datasets/{dataset_id}/versions/{second_version_id}",
        headers=auth_headers,
    )
    assert delete_resp.status_code == 200
    assert delete_resp.json() == {"deleted": True, "id": second_version_id}

    versions_after_resp = client.get(
        f"/api/datasets/{dataset_id}/versions",
        headers=auth_headers,
    )
    assert versions_after_resp.status_code == 200
    versions_after = versions_after_resp.json()
    assert len(versions_after) == 1
    assert versions_after[0]["id"] == active_version_id


def test_dataset_version_delete_blocks_active_version(
    client: TestClient,
    auth_headers,
    monkeypatch,
    tmp_path,
):
    dataset_id = _create_dataset_via_api(
        client,
        auth_headers,
        monkeypatch,
        tmp_path,
        _make_two_row_xlsx_bytes(),
        "Active Version Delete Project",
        "Active Version Delete Conn",
        "Active Version Delete Dataset",
    )

    versions_resp = client.get(f"/api/datasets/{dataset_id}/versions", headers=auth_headers)
    assert versions_resp.status_code == 200
    active_version_id = versions_resp.json()[0]["id"]

    version_summary_resp = client.get(
        f"/api/datasets/{dataset_id}/versions/{active_version_id}/dependencies",
        headers=auth_headers,
    )
    assert version_summary_resp.status_code == 200
    version_summary = version_summary_resp.json()
    assert version_summary["can_delete"] is False
    assert version_summary["is_active_version"] is True

    delete_resp = client.delete(
        f"/api/datasets/{dataset_id}/versions/{active_version_id}",
        headers=auth_headers,
    )
    assert delete_resp.status_code == 400
    assert "active version" in delete_resp.json()["detail"].lower()


def test_connection_soft_delete_hides_restore_and_permanent_delete(
    client: TestClient,
    auth_headers,
):
    _, connection_id = _create_project_and_connection(client, auth_headers, "Delete Project")

    soft_delete_resp = client.post(
        f"/api/connections/{connection_id}/soft-delete",
        headers=auth_headers,
    )
    assert soft_delete_resp.status_code == 200
    assert soft_delete_resp.json()["status"] == "soft_deleted"

    list_resp = client.get("/api/connections/", headers=auth_headers)
    assert list_resp.status_code == 200
    assert all(item["id"] != connection_id for item in list_resp.json())

    restore_resp = client.post(f"/api/connections/{connection_id}/restore", headers=auth_headers)
    assert restore_resp.status_code == 200
    assert restore_resp.json()["status"] == "active"

    second_soft_delete_resp = client.post(
        f"/api/connections/{connection_id}/soft-delete",
        headers=auth_headers,
    )
    assert second_soft_delete_resp.status_code == 200

    permanent_delete_resp = client.request(
        "DELETE",
        f"/api/connections/{connection_id}/permanent",
        headers=auth_headers,
    )
    assert permanent_delete_resp.status_code == 200
    assert permanent_delete_resp.json() == {"deleted": True, "id": connection_id}

    get_resp = client.get(f"/api/connections/{connection_id}", headers=auth_headers)
    assert get_resp.status_code == 404
