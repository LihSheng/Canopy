import pytest
from io import BytesIO

pytestmark = pytest.mark.api_schema


@pytest.mark.usefixtures("seed_analytics_data")
class TestExportsIntegration:
    def test_requires_auth(self, client):
        response = client.post("/api/exports/executive-summary", json={})
        assert response.status_code == 401

    def test_returns_excel_file(self, client, auth_headers):
        response = client.post(
            "/api/exports/executive-summary",
            json={"include_departments": True, "include_anomalies": True},
            headers=auth_headers,
        )
        assert response.status_code == 200
        assert response.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment" in response.headers["content-disposition"]
        assert response.content[:2] == b"PK"

    def test_excel_file_is_valid_workbook(self, client, auth_headers):
        response = client.post(
            "/api/exports/executive-summary",
            json={"include_departments": True, "include_anomalies": True},
            headers=auth_headers,
        )

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(response.content))
        assert "Executive Summary" in wb.sheetnames
        assert "Departments" in wb.sheetnames
        assert "Monthly Trends" in wb.sheetnames

    def test_excludes_departments_when_requested(self, client, auth_headers):
        response = client.post(
            "/api/exports/executive-summary",
            json={"include_departments": False, "include_anomalies": True},
            headers=auth_headers,
        )

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(response.content))
        assert "Departments" not in wb.sheetnames

    def test_excludes_anomalies_when_requested(self, client, auth_headers):
        response = client.post(
            "/api/exports/executive-summary",
            json={"include_departments": True, "include_anomalies": False},
            headers=auth_headers,
        )

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(response.content))
        assert "Anomalies" not in wb.sheetnames
