import io

import openpyxl
import pytest
from fastapi.testclient import TestClient


def _make_xlsx(rows: list[list]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _upload_xlsx(client: TestClient, auth_headers: dict, rows: list[list] | None = None) -> str:
    if rows is None:
        rows = [["name", "amount", "date"], ["Alice", 100, "2026-01-15"], ["Bob", 200, "2026-02-20"]]
    buf = _make_xlsx(rows)
    resp = client.post(
        "/api/v3/ingestion/uploads",
        files={"file": ("data.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"source_profile": "herdhr", "dataset_type": "payroll"},
        headers=auth_headers,
    )
    return resp.json()["upload_id"]


class TestPreviewEndpoint:
    def test_preview_returns_profile(self, client: TestClient, auth_headers):
        upload_id = _upload_xlsx(client, auth_headers)
        resp = client.get(f"/api/v3/ingestion/uploads/{upload_id}/preview", headers=auth_headers)
        assert resp.status_code == 200
        data = resp.json()
        assert data["upload_id"] == upload_id
        assert len(data["sheet_profiles"]) > 0
        assert len(data["column_profiles"]) > 0

    def test_preview_column_profiles(self, client: TestClient, auth_headers):
        upload_id = _upload_xlsx(client, auth_headers)
        resp = client.get(f"/api/v3/ingestion/uploads/{upload_id}/preview", headers=auth_headers)
        data = resp.json()
        col_names = [c["source_column_name"] for c in data["column_profiles"]]
        assert "name" in col_names
        assert "amount" in col_names
        assert "date" in col_names

    def test_preview_requires_auth(self, client: TestClient):
        resp = client.get("/api/v3/ingestion/uploads/fake-id/preview")
        assert resp.status_code == 401

    def test_preview_not_found(self, client: TestClient, auth_headers):
        resp = client.get("/api/v3/ingestion/uploads/nonexistent-id/preview", headers=auth_headers)
        assert resp.status_code == 404

    def test_preview_best_sheet_detected(self, client: TestClient, auth_headers):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["x", "y"])
        ws1.append([1, 2])
        ws2 = wb.create_sheet("Payroll")
        ws2.append(["Employee", "Salary", "Joined"])
        ws2.append(["Alice", 5000, "2025-01-01"])
        ws2.append(["Bob", 6000, "2025-03-15"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = client.post(
            "/api/v3/ingestion/uploads",
            files={"file": ("payroll.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"source_profile": "herdhr", "dataset_type": "payroll"},
            headers=auth_headers,
        )
        upload_id = resp.json()["upload_id"]

        preview = client.get(f"/api/v3/ingestion/uploads/{upload_id}/preview", headers=auth_headers)
        assert preview.status_code == 200
        data = preview.json()
        assert data["best_sheet_name"] == "Payroll"
        assert len(data["column_profiles"]) == 3
