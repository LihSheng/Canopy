from io import BytesIO

from openpyxl import load_workbook

from exports.builders.workbook import build_workbook
from exports.domain import (
    AnomalyExportRow,
    DepartmentExportRow,
    ExportPayload,
    MonthlyTrendExportRow,
)


def _make_payload(include_deps=True, include_anomalies=True, include_trends=True):
    return ExportPayload(
        snapshot_id="snap-001",
        generated_at="2026-05-15T10:00:00Z",
        departments=[
            DepartmentExportRow(
                rank=1,
                department_name="Engineering",
                total_spend=445000.00,
                payroll_spend=440000.00,
                claims_spend=5000.00,
                change_pct=2.30,
            ),
            DepartmentExportRow(
                rank=2,
                department_name="Sales",
                total_spend=328000.00,
                payroll_spend=325000.00,
                claims_spend=3000.00,
                change_pct=1.80,
            ),
        ]
        if include_deps
        else [],
        anomalies=[
            AnomalyExportRow(
                department_name="Engineering",
                period="2026-05",
                description="Total spend spike of 15.2% vs previous month",
                severity="high",
                change_pct=15.2,
            ),
        ]
        if include_anomalies
        else [],
        summary_payroll=1500000.00,
        summary_claims=45000.00,
        department_count=6,
        anomaly_count=3,
        period_label="2026-05",
        trends=[
            MonthlyTrendExportRow(month="2025-11", payroll=1424500.0, claims=32800.0, total=1457300.0),
            MonthlyTrendExportRow(month="2025-12", payroll=1430000.0, claims=33100.0, total=1463100.0),
        ]
        if include_trends
        else [],
    )


class TestWorkbookBuilder:
    def test_build_workbook_returns_bytes(self):
        payload = _make_payload()
        result = build_workbook(payload)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_output_is_valid_xlsx(self):
        payload = _make_payload()
        data = build_workbook(payload)
        wb = load_workbook(BytesIO(data))
        assert "Executive Summary" in wb.sheetnames
        assert "Departments" in wb.sheetnames
        assert "Anomalies" in wb.sheetnames
        assert "Monthly Trends" in wb.sheetnames

    def test_summary_sheet_contains_expected_values(self):
        payload = _make_payload()
        data = build_workbook(payload)
        wb = load_workbook(BytesIO(data))
        ws = wb["Executive Summary"]

        metric_names = [ws.cell(row=r, column=1).value for r in range(2, 10)]
        assert "Total Payroll (MYR)" in metric_names
        assert "Total Claims (MYR)" in metric_names
        assert "Department Count" in metric_names
        assert "Anomaly Count" in metric_names
        assert "Snapshot ID" in metric_names
        assert "Period" in metric_names

    def test_departments_sheet_contains_data(self):
        payload = _make_payload()
        data = build_workbook(payload)
        wb = load_workbook(BytesIO(data))
        ws = wb["Departments"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        assert "Rank" in headers
        assert "Department" in headers
        assert "Total Spend" in headers

        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=2, column=2).value == "Engineering"
        assert ws.cell(row=3, column=2).value == "Sales"

    def test_anomalies_sheet_contains_data(self):
        payload = _make_payload()
        data = build_workbook(payload)
        wb = load_workbook(BytesIO(data))
        ws = wb["Anomalies"]

        assert ws.cell(row=2, column=1).value == "Engineering"
        assert ws.cell(row=2, column=2).value == "2026-05"
        assert ws.cell(row=2, column=4).value == "high"

    def test_trends_sheet_contains_data(self):
        payload = _make_payload()
        data = build_workbook(payload)
        wb = load_workbook(BytesIO(data))
        ws = wb["Monthly Trends"]

        assert ws.cell(row=2, column=1).value == "2025-11"
        assert ws.cell(row=3, column=1).value == "2025-12"

    def test_no_departments_sheet_when_empty(self):
        payload = _make_payload(include_deps=False)
        data = build_workbook(payload)
        wb = load_workbook(BytesIO(data))
        assert "Departments" not in wb.sheetnames

    def test_no_anomalies_sheet_when_empty(self):
        payload = _make_payload(include_anomalies=False)
        data = build_workbook(payload)
        wb = load_workbook(BytesIO(data))
        assert "Anomalies" not in wb.sheetnames

    def test_no_trends_sheet_when_empty(self):
        payload = _make_payload(include_trends=False)
        data = build_workbook(payload)
        wb = load_workbook(BytesIO(data))
        assert "Monthly Trends" not in wb.sheetnames

    def test_money_format_applied_to_monetary_cells(self):
        payload = _make_payload()
        data = build_workbook(payload)
        wb = load_workbook(BytesIO(data))
        ws = wb["Departments"]

        cell = ws.cell(row=2, column=3)
        assert cell.number_format == "#,##0.00"

    def test_pct_format_applied_to_change_cells(self):
        payload = _make_payload()
        data = build_workbook(payload)
        wb = load_workbook(BytesIO(data))
        ws = wb["Departments"]

        cell = ws.cell(row=2, column=6)
        assert cell.number_format == '0.00"%"'
