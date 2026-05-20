import uuid
import io
from pathlib import Path

import pytest
import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from common.database import Base
from dataset.domain import Dataset, DatasetVersion, DatasetStatus, DatasetVersionStatus
from dataset.repository import DatasetRepository, DatasetVersionRepository
from dataset.service import DatasetVersionService
from dataset.preview_service import read_dataset_preview

@pytest.fixture(autouse=True)
def _setup_db():
    yield

def _make_sqlite_session():
    engine = create_engine("sqlite:///", connect_args={"check_same_thread": False})
    import dataset.schema  # noqa: F401
    Base.metadata.create_all(bind=engine)
    SessionLocal = sessionmaker(bind=engine)
    return SessionLocal()

class TestDatasetReimportService:
    def test_reimport_valid_data_creates_new_active_version(self, tmp_path):
        # Setup: Existing dataset with v1 active
        session = _make_sqlite_session()
        try:
            version_repo = DatasetVersionRepository(session)
            dataset_repo = DatasetRepository(session)
            service = DatasetVersionService(version_repo, dataset_repo)
            
            ds_id = str(uuid.uuid4())
            dataset = Dataset(id=ds_id, project_id="p1", connection_id="c1", name="test", status=DatasetStatus.ACTIVE.value)
            dataset_repo.save(dataset)
            
            v1 = DatasetVersion(id=str(uuid.uuid4()), dataset_id=ds_id, version_number=1, status=DatasetVersionStatus.READY.value)
            version_repo.save(v1)
            dataset_repo.update_active_version(ds_id, v1.id)

            data_path = tmp_path / "data.csv"
            data_path.write_text("A,B,C\n1,2,3\n", encoding="utf-8")
            
            # Action: Reimport
            new_version = service.reimport_version(
                ds_id,
                data_path=str(data_path),
                columns=["A", "B", "C"],
            )
            
            # Assert
            assert new_version.version_number == 2
            assert new_version.status == DatasetVersionStatus.READY.value
            assert new_version.row_count == 1
            assert new_version.column_count == 3
            assert Path(new_version.storage_path).exists()
            
            updated_dataset = dataset_repo.get(ds_id)
            assert updated_dataset.active_version_id == new_version.id
            
        finally:
            session.close()

    def test_reimport_materializes_previewable_storage(self, tmp_path):
        session = _make_sqlite_session()
        try:
            version_repo = DatasetVersionRepository(session)
            dataset_repo = DatasetRepository(session)
            service = DatasetVersionService(version_repo, dataset_repo)

            ds_id = str(uuid.uuid4())
            dataset = Dataset(
                id=ds_id,
                project_id="p1",
                connection_id="c1",
                name="payroll",
                source_object_name="Payroll",
                status=DatasetStatus.ACTIVE.value,
            )
            dataset_repo.save(dataset)

            csv_path = tmp_path / "payload.csv"
            csv_path.write_text("name,amount\nAlice,100\n", encoding="utf-8")

            new_version = service.reimport_version(
                ds_id,
                data_path=str(csv_path),
                columns=["name", "amount"],
            )

            assert new_version.status == DatasetVersionStatus.READY.value
            assert new_version.row_count == 1
            assert new_version.column_count == 2
            assert new_version.storage_path
            assert Path(new_version.storage_path).exists()

            preview = read_dataset_preview(new_version.storage_path)
            assert preview["columns"] == ["name", "amount"]
            assert preview["rows"] == [["Alice", "100"]]
            assert preview["total_row_count"] == 1
        finally:
            session.close()

    def test_reimport_uses_explicit_sheet_name_for_workbooks(self, tmp_path):
        session = _make_sqlite_session()
        try:
            version_repo = DatasetVersionRepository(session)
            dataset_repo = DatasetRepository(session)
            service = DatasetVersionService(version_repo, dataset_repo)

            ds_id = str(uuid.uuid4())
            dataset = Dataset(
                id=ds_id,
                project_id="p1",
                connection_id="c1",
                name="payroll",
                source_object_name="Payroll",
                status=DatasetStatus.ACTIVE.value,
            )
            dataset_repo.save(dataset)

            workbook = openpyxl.Workbook()
            first = workbook.active
            first.title = "Cover"
            first.append(["note"])
            payroll = workbook.create_sheet("Payroll")
            payroll.append(["name", "amount"])
            payroll.append(["Alice", 100])

            buffer = io.BytesIO()
            workbook.save(buffer)
            workbook.close()

            xlsx_path = tmp_path / "payload.xlsx"
            xlsx_path.write_bytes(buffer.getvalue())

            new_version = service.reimport_version(
                ds_id,
                data_path=str(xlsx_path),
                columns=["name", "amount"],
                sheet_name="Payroll",
            )

            assert new_version.row_count == 1
            assert new_version.column_count == 2
            preview = read_dataset_preview(new_version.storage_path)
            assert preview["columns"] == ["name", "amount"]
            assert preview["rows"] == [["Alice", 100]]
        finally:
            session.close()

    def test_reimport_falls_back_when_requested_sheet_is_missing(self, tmp_path):
        session = _make_sqlite_session()
        try:
            version_repo = DatasetVersionRepository(session)
            dataset_repo = DatasetRepository(session)
            service = DatasetVersionService(version_repo, dataset_repo)

            ds_id = str(uuid.uuid4())
            dataset = Dataset(
                id=ds_id,
                project_id="p1",
                connection_id="c1",
                name="payroll",
                source_object_name="Payroll",
                status=DatasetStatus.ACTIVE.value,
            )
            dataset_repo.save(dataset)

            workbook = openpyxl.Workbook()
            cover = workbook.active
            cover.title = "Cover"
            cover.append(["note"])
            payroll = workbook.create_sheet("Payroll")
            payroll.append(["name", "amount"])
            payroll.append(["Alice", 100])

            buffer = io.BytesIO()
            workbook.save(buffer)
            workbook.close()

            xlsx_path = tmp_path / "payload.xlsx"
            xlsx_path.write_bytes(buffer.getvalue())

            new_version = service.reimport_version(
                ds_id,
                data_path=str(xlsx_path),
                columns=["name", "amount"],
                sheet_name="MissingSheet",
            )

            assert new_version.row_count == 1
            assert new_version.column_count == 2
            preview = read_dataset_preview(new_version.storage_path)
            assert preview["columns"] == ["name", "amount"]
            assert preview["rows"] == [["Alice", 100]]
        finally:
            session.close()

    def test_mark_version_failed_sets_status_and_reason(self):
        session = _make_sqlite_session()
        try:
            version_repo = DatasetVersionRepository(session)
            dataset_repo = DatasetRepository(session)
            service = DatasetVersionService(version_repo, dataset_repo)

            version = DatasetVersion(
                id=str(uuid.uuid4()),
                dataset_id=str(uuid.uuid4()),
                version_number=1,
                status=DatasetVersionStatus.PENDING.value,
            )
            version_repo.save(version)

            result = service.mark_version_failed(version.id, "Schema mismatch: missing column A")
            assert result is not None
            assert result.status == DatasetVersionStatus.FAILED.value
            assert result.failure_reason == "Schema mismatch: missing column A"
        finally:
            session.close()

    def test_mark_version_failed_returns_none_for_missing(self):
        session = _make_sqlite_session()
        try:
            version_repo = DatasetVersionRepository(session)
            dataset_repo = DatasetRepository(session)
            service = DatasetVersionService(version_repo, dataset_repo)

            result = service.mark_version_failed("nonexistent-id", "reason")
            assert result is None
        finally:
            session.close()
