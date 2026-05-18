from pathlib import Path

import openpyxl
import pytest

from ingestion.sources.xlsx import (
    SheetData,
    WorkbookData,
    _propagate_merged_headers,
    _suggest_alternate_header,
    get_sheet_preview,
    get_workbook_metadata,
    read_workbook,
    sample_rows,
)


def _make_workbook(tmp_path: Path, sheets: dict) -> Path:
    path = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    for i, (name, data) in enumerate(sheets.items()):
        if i == 0:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(title=name)
        rows, kwargs = data.get("rows", []), data.get("options", {})
        for row in rows:
            ws.append(row)
        if not kwargs.get("visible", True):
            ws.sheet_state = "hidden"
        if kwargs.get("very_hidden"):
            ws.sheet_state = "veryHidden"
        for merge_range in kwargs.get("merge_ranges", []):
            ws.merge_cells(merge_range)
        for cell_ref, formula in kwargs.get("formulas", {}).items():
            ws[cell_ref] = formula
    wb.save(path)
    wb.close()
    return path


class TestSheetEnumeration:
    def test_enumerates_all_sheets(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Sheet1": {"rows": [["A"], [1]]},
            "Sheet2": {"rows": [["B"], [2]]},
            "Sheet3": {"rows": [["C"], [3]]},
        })
        result = read_workbook(path)
        assert len(result) == 3
        names = [s["sheet_name"] for s in result]
        assert names == ["Sheet1", "Sheet2", "Sheet3"]

    def test_column_count_matches(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Data": {"rows": [["A", "B", "C"], [1, 2, 3]]},
        })
        result = read_workbook(path)
        assert result[0]["column_count"] == 3

    def test_row_count_excludes_fully_empty_rows(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Data": {"rows": [["A"], [1], [None], [2], [None, None]]},
        })
        result = read_workbook(path)
        assert result[0]["row_count"] == 3


class TestHiddenSheets:
    def test_detects_hidden_sheet(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Visible": {"rows": [["A"], [1]]},
            "Hidden": {"rows": [["B"], [2]], "options": {"visible": False}},
        })
        result = read_workbook(path)
        assert result[0]["is_visible"] is True
        assert result[1]["is_visible"] is False

    def test_detects_very_hidden_sheet(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Visible": {"rows": [["A"], [1]]},
            "VeryHidden": {"rows": [["B"], [2]], "options": {"very_hidden": True}},
        })
        result = read_workbook(path)
        assert result[1]["is_visible"] is False

    def test_hidden_sheet_count_in_metadata(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "S1": {"rows": [["A"]]},
            "S2": {"rows": [["B"]], "options": {"visible": False}},
            "S3": {"rows": [["C"]], "options": {"visible": False}},
        })
        meta = get_workbook_metadata(path)
        assert meta["sheet_count"] == 3
        assert meta["visible_sheet_count"] == 1
        assert meta["hidden_sheet_count"] == 2


class TestEmptySheets:
    def test_empty_sheet_returns_no_rows(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Empty": {"rows": []},
        })
        result = read_workbook(path)
        assert result[0]["rows"] == []
        assert result[0]["row_count"] == 0

    def test_sheet_with_only_empty_rows(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Blanks": {"rows": [[None], [None, None]]},
        })
        result = read_workbook(path)
        assert result[0]["row_count"] == 0


class TestMergedCells:
    def test_detects_merged_cell_ranges(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Data": {
                "rows": [["Header", None], [1, 2]],
                "options": {"merge_ranges": ["A1:B1"]},
            },
        })
        result = read_workbook(path)
        assert "A1:B1" in result[0]["merged_cell_ranges"]

    def test_propagates_merged_header_values(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Data": {
                "rows": [["MergedHeader", None], [1, 2]],
                "options": {"merge_ranges": ["A1:B1"]},
            },
        })
        result = read_workbook(path)
        row0 = result[0]["rows"][0]
        assert row0[1] == "MergedHeader"

    def test_does_not_propagate_merged_data_rows(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Data": {
                "rows": [["H"], ["A", None]],
                "options": {"merge_ranges": ["A2:B2"]},
            },
        })
        result = read_workbook(path)
        row1 = result[0]["rows"][1]
        assert row1[1] is not None
        assert row1[1] == "A"


class TestFormulas:
    def test_detects_formula_cells(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Data": {
                "rows": [["A", "B"], [1, 2]],
                "options": {"formulas": {"A3": "=SUM(A1:A2)"}},
            },
        })
        result = read_workbook(path)
        assert result[0]["contains_formulas"] is True

    def test_no_formulas_flag_false(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Data": {"rows": [["A"], [1], [2]]},
        })
        result = read_workbook(path)
        assert result[0]["contains_formulas"] is False


class TestMultiRowHeaders:
    def test_suggests_alternate_header(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Data": {"rows": [["Category", "Details"], ["Name", "Age"], ["Alice", 30]]},
        })
        result = read_workbook(path)
        rows = result[0]["rows"]
        assert _suggest_alternate_header(rows) is True

    def test_no_alternate_header_single_row(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Data": {"rows": [["Name", "Age"], ["Alice", 30]]},
        })
        result = read_workbook(path)
        rows = result[0]["rows"]
        assert _suggest_alternate_header(rows) is False


class TestErrorHandling:
    def test_corrupted_workbook_returns_error(self, tmp_path):
        path = tmp_path / "corrupt.xlsx"
        path.write_bytes(b"not a valid excel file")
        result = read_workbook(path)
        assert len(result) == 0

    def test_metadata_contains_error_for_corrupted(self, tmp_path):
        path = tmp_path / "corrupt.xlsx"
        path.write_bytes(b"not a valid excel file")
        meta = get_workbook_metadata(path)
        assert "error" in meta

    def test_nonexistent_file_returns_error(self, tmp_path):
        path = tmp_path / "nonexistent.xlsx"
        meta = get_workbook_metadata(path)
        assert "error" in meta


class TestBackwardCompatibility:
    def test_read_workbook_returns_sheet_name_and_rows(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Data": {"rows": [["A", "B"], [1, 2]]},
        })
        result = read_workbook(path)
        assert "sheet_name" in result[0]
        assert "rows" in result[0]
        assert result[0]["sheet_name"] == "Data"
        assert len(result[0]["rows"]) == 2

    def test_new_fields_present(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Data": {"rows": [["A"], [1]]},
        })
        result = read_workbook(path)
        assert "is_visible" in result[0]
        assert "merged_cell_ranges" in result[0]
        assert "contains_formulas" in result[0]
        assert "warnings" in result[0]
        assert "column_count" in result[0]

    def test_sample_rows_unchanged(self):
        rows = [(1, 2), (3, 4), (5, 6)]
        assert sample_rows(rows, 2) == [(1, 2), (3, 4)]
        assert sample_rows([], 10) == []
        assert sample_rows(rows, 5) == rows
        assert sample_rows(rows, 3) == rows


class TestGetWorkbookMetadata:
    def test_returns_expected_keys(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "S1": {"rows": [["A"]]},
        })
        meta = get_workbook_metadata(path)
        assert "sheet_count" in meta
        assert "visible_sheet_count" in meta
        assert "hidden_sheet_count" in meta
        assert "contains_formulas" in meta

    def test_metadata_counts_correct(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "A": {"rows": [["A"]]},
            "B": {"rows": [["B"]], "options": {"visible": False}},
        })
        meta = get_workbook_metadata(path)
        assert meta["sheet_count"] == 2
        assert meta["visible_sheet_count"] == 1
        assert meta["hidden_sheet_count"] == 1


class TestGetSheetPreview:
    def test_returns_correct_number_of_rows(self, tmp_path):
        rows_data = [[i] for i in range(50)]
        path = _make_workbook(tmp_path, {
            "Data": {"rows": [["A"]] + rows_data},
        })
        preview = get_sheet_preview(path, "Data", max_rows=10)
        assert len(preview) == 10

    def test_returns_all_rows_when_fewer_than_max(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Data": {"rows": [["A"], [1], [2]]},
        })
        preview = get_sheet_preview(path, "Data", max_rows=20)
        assert len(preview) == 3

    def test_returns_empty_list_for_nonexistent_sheet(self, tmp_path):
        path = _make_workbook(tmp_path, {
            "Data": {"rows": [["A"]]},
        })
        preview = get_sheet_preview(path, "NoSuchSheet")
        assert preview == []

    def test_returns_empty_list_for_corrupted_file(self, tmp_path):
        path = tmp_path / "bad.xlsx"
        path.write_bytes(b"bad data")
        preview = get_sheet_preview(path, "Data")
        assert preview == []


class TestWorkbookDataDataclass:
    def test_workbook_data_defaults(self):
        wd = WorkbookData(sheets=[])
        assert wd.sheets == []
        assert wd.metadata == {}

    def test_sheet_data_defaults(self):
        sd = SheetData(sheet_name="Test", rows=[(1, 2)])
        assert sd.is_visible is True
        assert sd.row_count == 0
        assert sd.column_count == 0
        assert sd.merged_cell_ranges == []
        assert sd.contains_formulas is False
        assert sd.warnings == []
        assert sd.raw_cells is None


class TestPropagateMergedHeaders:
    def test_propagate_function_directly(self):
        rows = [("Header", None), (1, 2)]
        result = _propagate_merged_headers(rows, ["A1:B1"])
        assert result[0][1] == "Header"

    def test_propagate_empty_merged_ranges(self):
        rows = [("A", "B")]
        result = _propagate_merged_headers(rows, [])
        assert result == rows

    def test_propagate_no_rows(self):
        result = _propagate_merged_headers([], ["A1:B1"])
        assert result == []

