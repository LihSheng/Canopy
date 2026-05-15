from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from exports.builders.formatting import (
    apply_data_row_style,
    apply_header_style,
    apply_money_format,
    apply_pct_format,
    auto_width,
    write_title_row,
)


class TestFormattingHelpers:
    def test_apply_header_style_sets_font_fill_alignment_border(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="A")
        ws.cell(row=1, column=2, value="B")

        apply_header_style(ws, 1, 2)

        cell_a = ws.cell(row=1, column=1)
        cell_b = ws.cell(row=1, column=2)
        assert cell_a.font.bold is True
        assert cell_a.font.color.rgb == "00FFFFFF"
        assert cell_a.fill.start_color.rgb == "002F5496"
        assert cell_a.alignment.horizontal == "center"
        assert cell_b.font.bold is True

    def test_apply_data_row_style_sets_font_alignment_border(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=2, column=1, value="data")

        apply_data_row_style(ws, 2, 1)

        cell = ws.cell(row=2, column=1)
        assert cell.font.name == "Calibri"
        assert cell.alignment.horizontal == "center"

    def test_apply_data_row_alternate_applies_fill(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=3, column=1, value="alt")

        apply_data_row_style(ws, 3, 1, alternate=True)

        cell = ws.cell(row=3, column=1)
        assert cell.fill.start_color.rgb == "00D6E4F0"

    def test_apply_money_format(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=1234.56)

        apply_money_format(ws, 1, 1)

        assert ws.cell(row=1, column=1).number_format == '#,##0.00'

    def test_apply_pct_format(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=2, value=12.34)

        apply_pct_format(ws, 1, 2)

        assert ws.cell(row=1, column=2).number_format == '0.00"%"'

    def test_write_title_row(self):
        wb = Workbook()
        ws = wb.active

        next_row = write_title_row(ws, "My Title", 4)

        assert ws.cell(row=1, column=1).value == "My Title"
        assert ws.cell(row=1, column=1).font.size == 14
        assert next_row == 2

    def test_auto_width_sets_column_widths(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Short")
        ws.cell(row=2, column=1, value="A much longer value here")
        ws.cell(row=1, column=2, value="B")

        auto_width(ws, 2, min_width=8)

        assert ws.column_dimensions["A"].width >= 8
        assert ws.column_dimensions["B"].width >= 8

    def test_auto_width_respects_max(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="X" * 50)

        auto_width(ws, 1, max_width=20)

        assert ws.column_dimensions["A"].width == 20
