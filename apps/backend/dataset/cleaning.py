from __future__ import annotations

import copy
import csv
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from common.errors import ValidationError
from connection._shared import (
    is_empty_row,
    normalize_header,
    write_jsonl_version,
)


def clean_source_file(source_file_path: Path, sheet_name: str, dataset_id: str) -> dict:
    raw_rows = _read_rows(source_file_path, sheet_name)
    cleaned_result = clean_rows(raw_rows)

    effective_sheet_name = sheet_name if source_file_path.suffix.lower() == ".xlsx" else source_file_path.stem
    cleaned_path = write_jsonl_version(cleaned_result["cleaned_rows"], dataset_id, effective_sheet_name)

    return {
        "cleaned_path": str(cleaned_path),
        "raw_path": str(source_file_path),
        "row_count": cleaned_result["row_count"],
        "column_count": cleaned_result["column_count"],
        "cleaning_issues": cleaned_result["issues"],
        "columns": cleaned_result["columns"],
    }


def _read_rows(source_file_path: Path, sheet_name: str) -> list[dict]:
    if source_file_path.suffix.lower() == ".csv":
        return _read_csv_rows(source_file_path)
    return _read_xlsx_rows(source_file_path, sheet_name)


def _read_xlsx_rows(source_file_path: Path, sheet_name: str) -> list[dict]:
    workbook = load_workbook(filename=str(source_file_path), read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValidationError(f"Sheet '{sheet_name}' not found in workbook")
        worksheet = workbook[sheet_name]
        rows = [list(row) for row in worksheet.iter_rows(values_only=True) if not is_empty_row(row)]
        if not rows:
            return []
        header_row = rows[0]
        column_count = len(header_row)
        headers = normalize_header(header_row, column_count)
        return [
            {header: row[index] if index < len(row) else None for index, header in enumerate(headers)}
            for row in rows[1:]
        ]
    finally:
        workbook.close()


def _read_csv_rows(source_file_path: Path) -> list[dict]:
    with open(source_file_path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        rows = [row for row in reader if not is_empty_row(row)]
    if not rows:
        return []
    headers = normalize_header(rows[0], len(rows[0]))
    return [
        {header: row[index] if index < len(row) else None for index, header in enumerate(headers)} for row in rows[1:]
    ]


def clean_rows(rows: list[dict]) -> dict:
    working = copy.deepcopy(rows)
    issues: list[dict] = []

    working = _trim_whitespace(working, issues)
    working = _remove_empty_rows(working, issues)
    column_types = _infer_column_types(working)
    working, type_issues = _flag_invalid_cells(working, column_types)
    issues.extend(type_issues)

    normalized_headers = _normalize_headers(working)

    return {
        "cleaned_rows": [{normalized_headers.get(k, k): v for k, v in row.items()} for row in working],
        "columns": list(normalized_headers.values()),
        "column_types": {normalized_headers.get(k, k): v for k, v in column_types.items()},
        "row_count": len(working),
        "column_count": len(normalized_headers),
        "issues": issues,
    }


def _trim_whitespace(rows: list[dict], issues: list[dict]) -> list[dict]:
    for row_index, row in enumerate(rows):
        for key, value in list(row.items()):
            if isinstance(value, str):
                trimmed = value.strip()
                if trimmed != value:
                    row[key] = trimmed
    return rows


def _normalize_headers(rows: list[dict]) -> dict[str, str]:
    if not rows:
        return {}
    header_map: dict[str, str] = {}
    seen: set[str] = set()
    for key in rows[0]:
        normalized = _normalize_column_name(key)
        unique = normalized
        counter = 2
        while unique in seen:
            unique = f"{normalized}_{counter}"
            counter += 1
        seen.add(unique)
        header_map[key] = unique
    return header_map


def _normalize_column_name(name: str) -> str:
    cleaned = []
    for char in str(name).strip().lower():
        if char.isalnum():
            cleaned.append(char)
        elif cleaned and cleaned[-1] != "_":
            cleaned.append("_")
    result = "".join(cleaned).strip("_")
    return result or "column"


def _remove_empty_rows(rows: list[dict], issues: list[dict]) -> list[dict]:
    output: list[dict] = []
    for row_index, row in enumerate(rows):
        if _is_fully_empty(row):
            issues.append(
                {
                    "row_index": row_index,
                    "type": "empty_row",
                    "message": "Row is fully empty and was removed",
                }
            )
            continue
        output.append(row)
    return output


def _is_fully_empty(row: dict) -> bool:
    for value in row.values():
        if value not in (None, ""):
            return False
    return True


def _infer_column_types(rows: list[dict]) -> dict[str, str]:
    if not rows:
        return {}
    type_map: dict[str, str] = {}
    for key in rows[0]:
        type_map[key] = _infer_type_for_column(key, rows)
    return type_map


def _infer_type_for_column(key: str, rows: list[dict]) -> str:
    type_counts: dict[str, int] = {}
    for row in rows:
        value = row.get(key)
        vtype = _classify_value(value)
        type_counts[vtype] = type_counts.get(vtype, 0) + 1
    if not type_counts:
        return "text"
    best = max(type_counts, key=type_counts.get)
    return best if type_counts[best] > 0 else "text"


def _classify_value(value: object) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float, Decimal)):
        return "number"
    if isinstance(value, datetime):
        return "date"
    if isinstance(value, str):
        stripped = value.strip()
        if _looks_like_integer(stripped):
            return "number"
        if _looks_like_float(stripped):
            return "number"
        if _looks_like_date(stripped):
            return "date"
        if _looks_like_boolean(stripped):
            return "boolean"
        return "text"
    return "text"


def _looks_like_integer(value: str) -> bool:
    if not value:
        return False
    value = value.replace(",", "")
    if value.startswith("-"):
        value = value[1:]
    return value.isdigit()


def _looks_like_float(value: str) -> bool:
    if not value:
        return False
    value = value.replace(",", "")
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


def _looks_like_date(value: str) -> bool:
    if not value:
        return False
    formats = ["%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S", "%Y/%m/%d"]
    for fmt in formats:
        try:
            datetime.strptime(value, fmt)
            return True
        except ValueError:
            continue
    return False


def _looks_like_boolean(value: str) -> bool:
    return value.lower() in {"true", "false", "yes", "no", "1", "0"}


def _flag_invalid_cells(rows: list[dict], column_types: dict[str, str]) -> tuple[list[dict], list[dict]]:
    issues: list[dict] = []
    for row_index, row in enumerate(rows):
        for key, col_type in column_types.items():
            value = row.get(key)
            if value is None:
                continue
            if col_type == "number" and isinstance(value, str):
                stripped = value.strip()
                if not _looks_like_integer(stripped) and not _looks_like_float(stripped):
                    issues.append(
                        {
                            "row_index": row_index,
                            "column": key,
                            "value": value,
                            "expected_type": col_type,
                            "type": "invalid_cell",
                            "message": f"Value '{value}' is not a valid {col_type}",
                        }
                    )
            elif col_type == "date" and isinstance(value, str):
                stripped = value.strip()
                if not _looks_like_date(stripped):
                    issues.append(
                        {
                            "row_index": row_index,
                            "column": key,
                            "value": value,
                            "expected_type": col_type,
                            "type": "invalid_cell",
                            "message": f"Value '{value}' is not a valid date",
                        }
                    )
            elif col_type == "boolean" and isinstance(value, str):
                stripped = value.strip()
                if not _looks_like_boolean(stripped):
                    issues.append(
                        {
                            "row_index": row_index,
                            "column": key,
                            "value": value,
                            "expected_type": col_type,
                            "type": "invalid_cell",
                            "message": f"Value '{value}' is not a valid boolean",
                        }
                    )
    return rows, issues
