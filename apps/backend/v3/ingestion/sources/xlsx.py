from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.utils.exceptions import InvalidFileException


@dataclass
class SheetData:
    sheet_name: str
    rows: list[tuple]
    raw_cells: list[list] | None = None
    is_visible: bool = True
    row_count: int = 0
    column_count: int = 0
    merged_cell_ranges: list[str] = field(default_factory=list)
    contains_formulas: bool = False
    warnings: list[str] = field(default_factory=list)


@dataclass
class WorkbookData:
    sheets: list[SheetData]
    metadata: dict = field(default_factory=dict)


def _propagate_merged_headers(
    rows: list[tuple], merged_ranges: list[str], header_threshold: int = 2
) -> list[tuple]:
    if not rows or not merged_ranges:
        return rows
    result = list(rows)
    for mr_str in merged_ranges:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(mr_str)
        except (ValueError, IndexError):
            continue
        if max_row > header_threshold:
            continue
        row_idx = min_row - 1
        col_idx = min_col - 1
        if row_idx >= len(result) or col_idx >= len(result[row_idx]):
            continue
        top_left = result[row_idx][col_idx]
        if top_left is None:
            continue
        for r in range(min_row - 1, max_row):
            if r >= len(result):
                break
            row_list = list(result[r])
            for c in range(min_col - 1, max_col):
                if c < len(row_list) and row_list[c] is None:
                    row_list[c] = top_left
            result[r] = tuple(row_list)
    return result


def _detect_formulas(ws) -> bool:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                return True
    return False


def _suggest_alternate_header(rows: list[tuple]) -> bool:
    if len(rows) < 2:
        return False
    row0_strs = sum(1 for c in rows[0] if isinstance(c, str) and c.strip())
    row1_strs = sum(1 for c in rows[1] if isinstance(c, str) and c.strip())
    return row0_strs >= 2 and row1_strs >= 2


def _extract_workbook_data(path: str | Path) -> WorkbookData:
    path = Path(path)

    try:
        wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    except InvalidFileException:
        return WorkbookData(sheets=[], metadata={"error": "Invalid or corrupted workbook file"})
    except Exception as e:
        msg = str(e).lower()
        if any(kw in msg for kw in ("password", "protected", "encrypted")):
            return WorkbookData(sheets=[], metadata={"error": "Workbook is password-protected"})
        return WorkbookData(sheets=[], metadata={"error": f"Failed to open workbook: {e}"})

    total_sheets = len(wb.sheetnames)
    visible_count = 0
    hidden_count = 0
    any_formulas = False

    sheet_metas: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        is_visible = ws.sheet_state == "visible"
        if is_visible:
            visible_count += 1
        else:
            hidden_count += 1

        merged_ranges = [str(mr) for mr in ws.merged_cells.ranges]
        has_formulas = _detect_formulas(ws)
        if has_formulas:
            any_formulas = True
        col_count = ws.max_column or 0

        sheet_metas.append({
            "sheet_name": sheet_name,
            "is_visible": is_visible,
            "merged_cell_ranges": merged_ranges,
            "contains_formulas": has_formulas,
            "column_count": col_count,
        })

    wb.close()

    try:
        wb_data = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        sheets = [
            SheetData(
                sheet_name=sm["sheet_name"],
                rows=[],
                is_visible=sm["is_visible"],
                column_count=sm["column_count"],
                merged_cell_ranges=sm["merged_cell_ranges"],
                contains_formulas=sm["contains_formulas"],
                warnings=[f"Data rows unavailable: {e}"],
            )
            for sm in sheet_metas
        ]
        meta: dict[str, Any] = {
            "sheet_count": total_sheets,
            "visible_sheet_count": visible_count,
            "hidden_sheet_count": hidden_count,
            "contains_formulas": any_formulas,
            "error": str(e),
        }
        return WorkbookData(sheets=sheets, metadata=meta)

    sheet_datas: list[SheetData] = []
    for sm in sheet_metas:
        ws = wb_data[sm["sheet_name"]]
        raw_rows = list(ws.iter_rows(values_only=True))
        data_row_count = sum(1 for r in raw_rows if any(c is not None for c in r))
        propagated = _propagate_merged_headers(raw_rows, sm["merged_cell_ranges"])

        sheet_datas.append(SheetData(
            sheet_name=sm["sheet_name"],
            rows=propagated,
            is_visible=sm["is_visible"],
            row_count=data_row_count,
            column_count=sm["column_count"],
            merged_cell_ranges=sm["merged_cell_ranges"],
            contains_formulas=sm["contains_formulas"],
        ))

    wb_data.close()

    metadata: dict[str, Any] = {
        "sheet_count": total_sheets,
        "visible_sheet_count": visible_count,
        "hidden_sheet_count": hidden_count,
        "contains_formulas": any_formulas,
    }

    return WorkbookData(sheets=sheet_datas, metadata=metadata)


def read_workbook(path: str | Path) -> list[dict]:
    wd = _extract_workbook_data(path)
    result: list[dict] = []
    for sd in wd.sheets:
        entry: dict[str, Any] = {
            "sheet_name": sd.sheet_name,
            "rows": sd.rows,
            "is_visible": sd.is_visible,
            "row_count": sd.row_count,
            "column_count": sd.column_count,
            "merged_cell_ranges": sd.merged_cell_ranges,
            "contains_formulas": sd.contains_formulas,
            "warnings": sd.warnings,
        }
        result.append(entry)
    return result


def sample_rows(rows: list, max_samples: int = 20) -> list:
    if not rows:
        return []
    if len(rows) <= max_samples:
        return rows
    return rows[:max_samples]


def get_workbook_metadata(path: str | Path) -> dict:
    wd = _extract_workbook_data(path)
    return wd.metadata


def get_sheet_preview(path: str | Path, sheet_name: str, max_rows: int = 20) -> list[tuple]:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    count = 0
    rows: list[tuple] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(row)
        count += 1
        if count >= max_rows:
            break
    wb.close()
    return rows
