from __future__ import annotations

import csv
import json
import uuid
from collections.abc import Iterable
from pathlib import Path

from fastapi import UploadFile
from openpyxl import load_workbook

from common.config import settings
from common.errors import ValidationError


def _storage_root() -> Path:
    if settings.export_storage_dir:
        return Path(settings.export_storage_dir) / "data-sources"
    return Path(__file__).resolve().parents[2] / "storage" / "data-sources"


def _is_empty_row(values: Iterable[object | None]) -> bool:
    for value in values:
        if value not in (None, ""):
            return False
    return True


def _normalize_header(values: list[object | None], column_count: int) -> list[str]:
    headers: list[str] = []
    for index in range(column_count):
        value = values[index] if index < len(values) else None
        if value in (None, ""):
            headers.append(f"column_{index + 1}")
        else:
            headers.append(str(value))
    return headers


def save_uploaded_file(file: UploadFile) -> Path:
    if file.filename is None:
        raise ValidationError("File name is required")

    suffix = Path(file.filename).suffix.lower()
    if suffix not in {".xlsx", ".csv"}:
        raise ValidationError("Only .xlsx and .csv files are supported for static file sources")

    storage_dir = _storage_root() / str(uuid.uuid4())
    storage_dir.mkdir(parents=True, exist_ok=True)

    storage_path = storage_dir / file.filename
    file_bytes = file.file.read()
    storage_path.write_bytes(file_bytes)
    return storage_path


def build_sheet_profiles(storage_path: Path) -> list[dict]:
    if storage_path.suffix.lower() == ".csv":
        return [_build_csv_profile(storage_path)]

    workbook = load_workbook(filename=str(storage_path), read_only=True, data_only=True)
    try:
        profiles: list[dict] = []
        for worksheet in workbook.worksheets:
            row_count = 0
            column_count = 0
            header_row_index = None

            for index, row in enumerate(worksheet.iter_rows(values_only=True)):
                if _is_empty_row(row):
                    continue
                row_count += 1
                column_count = max(column_count, len(row))
                if header_row_index is None:
                    header_row_index = index

            profiles.append(
                {
                    "sheet_name": worksheet.title,
                    "row_count": row_count,
                    "column_count": column_count,
                    "header_row_index": header_row_index,
                    "confidence": 1.0 if row_count > 0 else 0.0,
                    "warnings": [] if row_count > 0 else ["Empty sheet"],
                },
            )

        return profiles
    finally:
        workbook.close()


def materialize_dataset_version(storage_path: Path, sheet_name: str, dataset_id: str) -> tuple[Path, int, int]:
    if storage_path.suffix.lower() == ".csv":
        return _materialize_csv_version(storage_path, dataset_id)

    workbook = load_workbook(filename=str(storage_path), read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValidationError(f"Sheet '{sheet_name}' not found in workbook")

        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        non_empty_rows = [list(row) for row in rows if not _is_empty_row(row)]
        if not non_empty_rows:
            return _write_jsonl_version([], dataset_id, sheet_name), 0, 0

        header_row = non_empty_rows[0]
        column_count = len(header_row)
        headers = _normalize_header(header_row, column_count)
        data_rows = non_empty_rows[1:]
        output_rows = []
        for row in data_rows:
            row_values = list(row)
            output_rows.append({header: row_values[index] if index < len(row_values) else None for index, header in enumerate(headers)})

        return _write_jsonl_version(output_rows, dataset_id, sheet_name), len(output_rows), column_count
    finally:
        workbook.close()


def _build_csv_profile(storage_path: Path) -> dict:
    with open(storage_path, "r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        rows = [row for row in reader if not _is_empty_row(row)]

    if not rows:
        return {
            "sheet_name": storage_path.stem,
            "row_count": 0,
            "column_count": 0,
            "header_row_index": None,
            "confidence": 0.0,
            "warnings": ["Empty file"],
        }

    return {
        "sheet_name": storage_path.stem,
        "row_count": len(rows),
        "column_count": len(rows[0]),
        "header_row_index": 0,
        "confidence": 1.0,
        "warnings": [],
    }


def _materialize_csv_version(storage_path: Path, dataset_id: str) -> tuple[Path, int, int]:
    with open(storage_path, "r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        rows = [row for row in reader if not _is_empty_row(row)]

    if not rows:
        return _write_jsonl_version([], dataset_id, storage_path.stem), 0, 0

    headers = _normalize_header(rows[0], len(rows[0]))
    output_rows = []
    for row in rows[1:]:
        output_rows.append({header: row[index] if index < len(row) else None for index, header in enumerate(headers)})

    return _write_jsonl_version(output_rows, dataset_id, storage_path.stem), len(output_rows), len(headers)


def _write_jsonl_version(rows: list[dict], dataset_id: str, sheet_name: str) -> Path:
    version_dir = _storage_root() / dataset_id
    version_dir.mkdir(parents=True, exist_ok=True)
    version_path = version_dir / f"{_slugify(sheet_name)}.jsonl"
    with open(version_path, "w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, default=str))
            handle.write("\n")
    return version_path


def _slugify(value: str) -> str:
    cleaned = []
    for char in value.lower():
        if char.isalnum():
            cleaned.append(char)
        elif cleaned and cleaned[-1] != "-":
            cleaned.append("-")
    slug = "".join(cleaned).strip("-")
    return slug or "sheet"
