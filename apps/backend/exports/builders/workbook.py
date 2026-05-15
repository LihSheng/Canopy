from io import BytesIO

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from exports.builders.formatting import (
    apply_data_row_style,
    apply_header_style,
    apply_money_format,
    apply_pct_format,
    auto_width,
    write_title_row,
)
from exports.domain import ExportPayload


def build_workbook(payload: ExportPayload) -> bytes:
    wb = Workbook()

    _build_summary_sheet(wb, payload)

    if payload.departments:
        _build_departments_sheet(wb, payload.departments)

    if payload.anomalies:
        _build_anomalies_sheet(wb, payload.anomalies)

    if payload.trends:
        _build_trends_sheet(wb, payload.trends)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _build_summary_sheet(wb: Workbook, payload: ExportPayload) -> None:
    ws = wb.active
    ws.title = "Executive Summary"

    headers = ["Metric", "Value"]
    _write_headers(ws, 1, headers)
    apply_header_style(ws, 1, len(headers))

    rows = [
        ("Snapshot ID", payload.snapshot_id),
        ("Period", payload.period_label),
        ("Generated At", payload.generated_at),
        ("Total Payroll (MYR)", payload.summary_payroll),
        ("Total Claims (MYR)", payload.summary_claims),
        ("Total Spend (MYR)", payload.summary_payroll + payload.summary_claims),
        ("Department Count", payload.department_count),
        ("Anomaly Count", payload.anomaly_count),
    ]

    for i, (metric, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=metric)
        ws.cell(row=i, column=2, value=value)
        apply_data_row_style(ws, i, len(headers), alternate=(i % 2 == 0))
        if metric.startswith("Total"):
            apply_money_format(ws, i, 2)

    auto_width(ws, len(headers))


def _build_departments_sheet(
    wb: Workbook, departments: list
) -> None:
    ws = wb.create_sheet(title="Departments")
    headers = ["Rank", "Department", "Total Spend", "Payroll", "Claims", "Change %"]
    _write_headers(ws, 1, headers)
    apply_header_style(ws, 1, len(headers))

    for i, dept in enumerate(departments, start=2):
        ws.cell(row=i, column=1, value=dept.rank)
        ws.cell(row=i, column=2, value=dept.department_name)
        ws.cell(row=i, column=3, value=dept.total_spend)
        ws.cell(row=i, column=4, value=dept.payroll_spend)
        ws.cell(row=i, column=5, value=dept.claims_spend)
        ws.cell(row=i, column=6, value=dept.change_pct)
        apply_data_row_style(ws, i, len(headers), alternate=(i % 2 == 0))
        for col in (3, 4, 5):
            apply_money_format(ws, i, col)
        apply_pct_format(ws, i, 6)

    auto_width(ws, len(headers))


def _build_anomalies_sheet(
    wb: Workbook, anomalies: list
) -> None:
    ws = wb.create_sheet(title="Anomalies")
    headers = ["Department", "Period", "Description", "Severity", "Change %"]
    _write_headers(ws, 1, headers)
    apply_header_style(ws, 1, len(headers))

    for i, anom in enumerate(anomalies, start=2):
        ws.cell(row=i, column=1, value=anom.department_name)
        ws.cell(row=i, column=2, value=anom.period)
        ws.cell(row=i, column=3, value=anom.description)
        ws.cell(row=i, column=4, value=anom.severity)
        ws.cell(row=i, column=5, value=anom.change_pct)
        apply_data_row_style(ws, i, len(headers), alternate=(i % 2 == 0))
        apply_pct_format(ws, i, 5)

    auto_width(ws, len(headers))


def _build_trends_sheet(
    wb: Workbook, trends: list
) -> None:
    ws = wb.create_sheet(title="Monthly Trends")
    headers = ["Month", "Payroll", "Claims", "Total"]
    _write_headers(ws, 1, headers)
    apply_header_style(ws, 1, len(headers))

    for i, trend in enumerate(trends, start=2):
        ws.cell(row=i, column=1, value=trend.month)
        ws.cell(row=i, column=2, value=trend.payroll)
        ws.cell(row=i, column=3, value=trend.claims)
        ws.cell(row=i, column=4, value=trend.total)
        apply_data_row_style(ws, i, len(headers), alternate=(i % 2 == 0))
        for col in (2, 3, 4):
            apply_money_format(ws, i, col)

    auto_width(ws, len(headers))


def _write_headers(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col_idx, value=header)
