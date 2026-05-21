from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

_TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F3864")
_SUBTITLE_FONT = Font(name="Calibri", size=11, bold=True, color="2F5496")

_DATA_FONT = Font(name="Calibri", size=11)
_MONEY_FORMAT = "#,##0.00"
_PCT_FORMAT = '0.00"%"'

_THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

_ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def apply_header_style(ws: Worksheet, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER


def apply_data_row_style(ws: Worksheet, row: int, col_count: int, alternate: bool = False) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _DATA_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER
        if alternate:
            cell.fill = _ALT_FILL


def apply_money_format(ws: Worksheet, row: int, col: int) -> None:
    ws.cell(row=row, column=col).number_format = _MONEY_FORMAT


def apply_pct_format(ws: Worksheet, row: int, col: int) -> None:
    ws.cell(row=row, column=col).number_format = _PCT_FORMAT


def write_title_row(ws: Worksheet, title: str, col_count: int) -> int:
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = _TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    return 2


def auto_width(ws: Worksheet, col_count: int, min_width: int = 12, max_width: int = 36) -> None:
    for col_idx in range(1, col_count + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = min_width
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            for cell_value in row:
                if cell_value is not None:
                    max_len = max(max_len, min(len(str(cell_value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len
